"""Tests for `tools/render_workbooks.py`.

Covers lazy creation, mtime-skip cache, privacy scan posture, opt-in
gating (health + business), per-entity workbooks (pet / vehicle /
asset), and the 13 default-domain dispatchers.
"""
from __future__ import annotations

import time
from pathlib import Path

import yaml
from openpyxl import load_workbook

from superagent.tools.render_workbooks import (
    DOMAIN_RENDERERS,
    ENTITY_RENDERERS,
    REDACTED,
    render_all,
    render_domain,
    render_entity,
    workbook_config,
)


def _write(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w") as fh:
        yaml.safe_dump(data, fh, sort_keys=False)


def _seed_bills(workspace: Path, rows: list[dict]) -> None:
    _write(workspace / "_memory" / "bills.yaml", {"schema_version": 1, "bills": rows})


def _seed_assets(workspace: Path, rows: list[dict]) -> None:
    _write(
        workspace / "_memory" / "assets-index.yaml",
        {"schema_version": 1, "assets": rows},
    )


def _seed_accounts(workspace: Path, rows: list[dict]) -> None:
    _write(
        workspace / "_memory" / "accounts-index.yaml",
        {"schema_version": 1, "accounts": rows},
    )


def _seed_contacts(workspace: Path, rows: list[dict]) -> None:
    _write(
        workspace / "_memory" / "contacts.yaml",
        {"schema_version": 1, "contacts": rows},
    )


def _seed_appointments(workspace: Path, rows: list[dict]) -> None:
    _write(
        workspace / "_memory" / "appointments.yaml",
        {"schema_version": 1, "appointments": rows},
    )


# ---------------------------------------------------------------------------
# Dispatcher coverage
# ---------------------------------------------------------------------------

def test_all_thirteen_default_domains_have_renderers() -> None:
    assert sorted(DOMAIN_RENDERERS) == sorted([
        "assets", "business", "career", "education", "family",
        "finances", "health", "hobbies", "home", "pets",
        "self", "travel", "vehicles",
    ])


def test_three_entity_kinds_have_renderers() -> None:
    assert sorted(ENTITY_RENDERERS) == ["asset", "pet", "vehicle"]


# ---------------------------------------------------------------------------
# Lazy creation
# ---------------------------------------------------------------------------

def test_render_empty_workspace_skips_all_workbooks(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    results = render_all(initialized_workspace, framework_dir)
    assert all(r.status.startswith("skipped") for r in results), \
        f"unexpected non-skipped results: {[r.status for r in results]}"
    domains_root = initialized_workspace / "Domains"
    if domains_root.exists():
        assert not list(domains_root.rglob("*.xlsx")), (
            "no workbooks should be written for an empty workspace"
        )


def test_render_finances_skips_when_no_data(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    result = render_domain(initialized_workspace, framework_dir, "finances")
    assert result.status == "skipped-empty"
    assert result.path is None


def test_render_finances_emits_workbook_when_bills_exist(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_bills(initialized_workspace, [
        {"id": "pge", "name": "PG&E", "kind": "utility", "amount": 120,
         "cadence": "monthly", "next_due": "2026-06-01"},
    ])
    result = render_domain(initialized_workspace, framework_dir, "finances")
    assert result.status == "rendered"
    assert result.path is not None
    assert result.path.exists()
    assert "Bills" in result.sheets

    wb = load_workbook(result.path)
    assert "Bills" in wb.sheetnames
    ws = wb["Bills"]
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=2, column=1).value == "PG&E"
    assert ws.freeze_panes == "A2"


def test_render_assets_includes_financial_holding_sheet(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_assets(initialized_workspace, [
        {"id": "aapl", "name": "AAPL holdings", "kind": "stock",
         "domain": "assets", "ticker": "AAPL", "exchange": "NASDAQ",
         "units": 100, "cost_basis": 15000, "current_value": 18000,
         "held_in_account": "account:schwab-brokerage", "status": "active"},
    ])
    result = render_domain(initialized_workspace, framework_dir, "assets")
    assert result.status == "rendered"
    wb = load_workbook(result.path)
    assert "Financial Holdings" in wb.sheetnames
    ws = wb["Financial Holdings"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(ws[1]) + 1)]
    assert "Ticker" in headers
    assert "Held in account" in headers


# ---------------------------------------------------------------------------
# Mtime skip cache
# ---------------------------------------------------------------------------

def test_re_render_is_skipped_when_sources_unchanged(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_bills(initialized_workspace, [
        {"id": "x", "name": "X", "kind": "utility", "amount": 1,
         "cadence": "monthly"},
    ])
    first = render_domain(initialized_workspace, framework_dir, "finances")
    assert first.status == "rendered"
    second = render_domain(initialized_workspace, framework_dir, "finances")
    assert second.status == "skipped-stale"


def test_re_render_fires_when_source_mtime_advances(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_bills(initialized_workspace, [
        {"id": "x", "name": "X", "kind": "utility", "amount": 1,
         "cadence": "monthly"},
    ])
    render_domain(initialized_workspace, framework_dir, "finances")
    time.sleep(1.1)
    _seed_bills(initialized_workspace, [
        {"id": "x", "name": "X", "kind": "utility", "amount": 2,
         "cadence": "monthly"},
        {"id": "y", "name": "Y", "kind": "utility", "amount": 3,
         "cadence": "monthly"},
    ])
    third = render_domain(initialized_workspace, framework_dir, "finances")
    assert third.status == "rendered"


def test_re_render_fires_when_history_window_changes(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_bills(initialized_workspace, [
        {"id": "x", "name": "X", "kind": "utility", "amount": 1,
         "cadence": "monthly"},
    ])
    render_domain(initialized_workspace, framework_dir, "finances")

    config_path = initialized_workspace / "_memory" / "config.yaml"
    cfg = yaml.safe_load(config_path.read_text())
    cfg.setdefault("preferences", {}).setdefault("workbooks", {})
    cfg["preferences"]["workbooks"]["history_window_years"] = 5
    with config_path.open("w") as fh:
        yaml.safe_dump(cfg, fh, sort_keys=False)

    third = render_domain(initialized_workspace, framework_dir, "finances")
    assert third.status == "rendered"


def test_skipped_empty_removes_stale_workbook(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_bills(initialized_workspace, [
        {"id": "x", "name": "X", "kind": "utility", "amount": 1,
         "cadence": "monthly"},
    ])
    first = render_domain(initialized_workspace, framework_dir, "finances")
    assert first.status == "rendered" and first.path is not None
    workbook_path = first.path

    # Empty out source data — workbook should be removed on next render.
    _seed_bills(initialized_workspace, [])
    _write(initialized_workspace / "_memory" / "subscriptions.yaml",
           {"schema_version": 1, "subscriptions": []})
    _write(initialized_workspace / "_memory" / "accounts-index.yaml",
           {"schema_version": 1, "accounts": []})
    second = render_domain(initialized_workspace, framework_dir, "finances")
    assert second.status == "skipped-empty"
    assert not workbook_path.exists()


# ---------------------------------------------------------------------------
# Privacy scan
# ---------------------------------------------------------------------------

def test_privacy_scan_redacts_ssn_in_strict_mode(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_bills(initialized_workspace, [
        {"id": "x", "name": "X", "kind": "utility", "amount": 1,
         "cadence": "monthly", "tags": ["ssn-123-45-6789"]},
    ])
    result = render_domain(initialized_workspace, framework_dir, "finances")
    assert result.status == "rendered"
    wb = load_workbook(result.path)
    cells = list(wb["Bills"].iter_rows(values_only=True))
    flat = " ".join(str(c) for row in cells for c in row if c)
    assert "123-45-6789" not in flat
    assert REDACTED in flat
    assert result.redactions >= 1


def test_privacy_scan_off_lets_secrets_through(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    config_path = initialized_workspace / "_memory" / "config.yaml"
    cfg = yaml.safe_load(config_path.read_text())
    cfg.setdefault("preferences", {}).setdefault("workbooks", {})
    cfg["preferences"]["workbooks"]["privacy_scan"] = "off"
    with config_path.open("w") as fh:
        yaml.safe_dump(cfg, fh, sort_keys=False)

    _seed_bills(initialized_workspace, [
        {"id": "x", "name": "123-45-6789", "kind": "utility", "amount": 1,
         "cadence": "monthly"},
    ])
    result = render_domain(initialized_workspace, framework_dir, "finances")
    assert result.status == "rendered"
    wb = load_workbook(result.path)
    cells = list(wb["Bills"].iter_rows(values_only=True))
    flat = " ".join(str(c) for row in cells for c in row if c)
    assert "123-45-6789" in flat
    assert result.redactions == 0


# ---------------------------------------------------------------------------
# Opt-in (health + business)
# ---------------------------------------------------------------------------

def test_health_workbook_skips_when_opt_in_disabled(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _write(
        initialized_workspace / "_memory" / "health-records.yaml",
        {"schema_version": 1, "vitals": [
            {"date": "2026-05-01", "type": "weight", "value": 170,
             "unit": "lb", "source": "manual"},
        ]},
    )
    result = render_domain(initialized_workspace, framework_dir, "health")
    assert result.status == "skipped-empty"  # config opt-in is OFF


def test_health_workbook_renders_when_opt_in_enabled(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    config_path = initialized_workspace / "_memory" / "config.yaml"
    cfg = yaml.safe_load(config_path.read_text())
    cfg.setdefault("preferences", {}).setdefault("workbooks", {})
    cfg["preferences"]["workbooks"]["health"] = {"enabled": True}
    with config_path.open("w") as fh:
        yaml.safe_dump(cfg, fh, sort_keys=False)

    _write(
        initialized_workspace / "_memory" / "health-records.yaml",
        {"schema_version": 1, "vitals": [
            {"date": "2026-05-01", "type": "weight", "value": 170,
             "unit": "lb", "source": "manual"},
        ]},
    )
    result = render_domain(initialized_workspace, framework_dir, "health")
    assert result.status == "rendered"
    wb = load_workbook(result.path)
    assert "Vitals" in wb.sheetnames


def test_business_workbook_skips_when_opt_in_disabled(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_contacts(initialized_workspace, [
        {"id": "client-a", "name": "Client A",
         "related_domains": ["business"]},
    ])
    result = render_domain(initialized_workspace, framework_dir, "business")
    assert result.status == "skipped-empty"


# ---------------------------------------------------------------------------
# Per-entity workbooks
# ---------------------------------------------------------------------------

def test_entity_below_threshold_skips(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_appointments(initialized_workspace, [
        {"id": "a1", "kind": "vet", "for_member": "buddy",
         "date": "2026-05-01"},
        {"id": "a2", "kind": "vet", "for_member": "buddy",
         "date": "2026-05-02"},
    ])
    result = render_entity(initialized_workspace, framework_dir, "pet", "buddy")
    assert result.status == "skipped-empty"  # below DEFAULT_PER_ENTITY_MIN_EVENTS=5


def test_pet_entity_renders_when_event_count_meets_threshold(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_appointments(initialized_workspace, [
        {"id": f"a{i}", "kind": "vet", "for_member": "buddy",
         "date": f"2026-04-{i:02d}", "provider": "Dr. Vet",
         "reason": "checkup"}
        for i in range(1, 7)  # 6 visits, above threshold
    ])
    result = render_entity(initialized_workspace, framework_dir, "pet", "buddy")
    assert result.status == "rendered"
    wb = load_workbook(result.path)
    assert "Vet Visits" in wb.sheetnames


def test_per_entity_disabled_via_config(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    config_path = initialized_workspace / "_memory" / "config.yaml"
    cfg = yaml.safe_load(config_path.read_text())
    cfg.setdefault("preferences", {}).setdefault("workbooks", {})
    cfg["preferences"]["workbooks"]["per_entity"] = False
    with config_path.open("w") as fh:
        yaml.safe_dump(cfg, fh, sort_keys=False)

    _seed_appointments(initialized_workspace, [
        {"id": f"a{i}", "kind": "vet", "for_member": "buddy",
         "date": f"2026-04-{i:02d}"} for i in range(1, 7)
    ])
    result = render_entity(initialized_workspace, framework_dir, "pet", "buddy")
    assert result.status == "skipped-disabled"


def test_unknown_entity_kind_returns_error(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    result = render_entity(initialized_workspace, framework_dir, "fish", "x")
    assert result.status == "error"
    assert "unknown entity kind" in result.error


# ---------------------------------------------------------------------------
# Domain folder materialization
# ---------------------------------------------------------------------------

def test_render_materializes_domain_folder_lazily(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    """The lazy-domain-folder contract is honored: rendering a workbook
    creates `Domains/<Name>/` if it didn't exist."""
    domain_folder = initialized_workspace / "Domains" / "Finances"
    assert not domain_folder.exists()
    _seed_bills(initialized_workspace, [
        {"id": "x", "name": "X", "kind": "utility", "amount": 1,
         "cadence": "monthly"},
    ])
    render_domain(initialized_workspace, framework_dir, "finances")
    assert domain_folder.is_dir()
    assert (domain_folder / "finances.xlsx").is_file()


# ---------------------------------------------------------------------------
# Config helpers
# ---------------------------------------------------------------------------

def test_workbook_config_defaults_when_missing(
    initialized_workspace: Path,
) -> None:
    cfg = workbook_config(initialized_workspace)
    assert cfg["enabled"] is True
    assert cfg["history_window_years"] >= 1
    assert cfg["per_entity"] in (True, False)
    assert cfg["health"]["enabled"] is False
    assert cfg["business"]["enabled"] is False


def test_unknown_domain_returns_error(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    result = render_domain(initialized_workspace, framework_dir, "xyz")
    assert result.status == "error"


# ---------------------------------------------------------------------------
# Check / dry-run
# ---------------------------------------------------------------------------

def test_check_mode_does_not_write_file(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    _seed_bills(initialized_workspace, [
        {"id": "x", "name": "X", "kind": "utility", "amount": 1,
         "cadence": "monthly"},
    ])
    result = render_domain(initialized_workspace, framework_dir, "finances",
                            check=True)
    assert result.status == "rendered"
    assert result.path is not None
    assert not result.path.exists()


# ---------------------------------------------------------------------------
# render_all sanity
# ---------------------------------------------------------------------------

def test_render_all_handles_mixed_populated_and_empty(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    """A non-utility bill seeded → only finances renders; everything else
    (no source data) is skipped-empty.
    """
    _seed_bills(initialized_workspace, [
        {"id": "comcast", "name": "Comcast Internet", "kind": "internet",
         "amount": 89, "cadence": "monthly"},
    ])
    results = render_all(initialized_workspace, framework_dir)
    by_status: dict[str, list[str]] = {}
    for r in results:
        by_status.setdefault(r.status, []).append(r.target)
    assert by_status.get("rendered") == ["domain:finances"], (
        f"only finances should render; got {by_status.get('rendered')}"
    )
    assert "skipped-empty" in by_status
    assert "domain:assets" in by_status["skipped-empty"]
    assert "domain:education" in by_status["skipped-empty"]


def test_render_all_cross_renders_when_data_implicates_multiple_domains(
    framework_dir: Path,
    initialized_workspace: Path,
) -> None:
    """A utility bill seeded → BOTH finances (Bills sheet) and home
    (Utilities sheet) render. This is the correct cross-rendering
    behavior — utilities are home-domain by convention.
    """
    _seed_bills(initialized_workspace, [
        {"id": "pge", "name": "PG&E", "kind": "utility", "amount": 120,
         "cadence": "monthly"},
    ])
    results = render_all(initialized_workspace, framework_dir)
    rendered = {r.target for r in results if r.status == "rendered"}
    assert rendered == {"domain:finances", "domain:home"}

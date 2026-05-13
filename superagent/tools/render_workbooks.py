#!/usr/bin/env -S uv run python
"""Lazy per-domain (and per-entity) `.xlsx` workbook projections.

Implements the design in `workspace/Outbox/drafts/2026-05-12-tabular-projections-proposal.md`
(approved 2026-05-13). Source-of-truth stays in YAML / markdown; the
workbook is a human-friendly read-only projection (the agent never reads
it back).

CLI:

    uv run python -m superagent.tools.render_workbooks --domain <id>
    uv run python -m superagent.tools.render_workbooks --entity <kind>:<slug>
    uv run python -m superagent.tools.render_workbooks --all
    uv run python -m superagent.tools.render_workbooks --status
    uv run python -m superagent.tools.render_workbooks --enable health|business
    uv run python -m superagent.tools.render_workbooks --disable health|business
    uv run python -m superagent.tools.render_workbooks --check        # dry-run

Lazy rules:
  - A workbook is emitted only when ≥1 sheet has ≥1 data row.
  - A re-render is skipped when every source-file mtime is unchanged AND
    the active config (history_window_years, privacy posture) is unchanged.
  - The `health` and `business` workbooks are opt-in (default OFF) per the
    sensitivity matrix in the proposal.

Per-entity workbooks (when `config.preferences.workbooks.per_entity: true`)
materialize alongside the per-domain workbook for high-event entities:
  - `Domains/Pets/<pet-slug>.xlsx`        — when ≥1 logged event for the pet.
  - `Domains/Vehicles/<vehicle-slug>.xlsx` — when ≥1 vehicle-log event.
  - `Domains/Assets/<asset-slug>.xlsx`     — when financial holding has
                                              ≥2 lots OR a transaction history.
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import re
import sys
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Module constants
# ---------------------------------------------------------------------------

DEFAULT_HISTORY_YEARS = 10
DEFAULT_PER_ENTITY_MIN_EVENTS = 5     # per-entity workbook threshold
DEFAULT_PER_ENTITY_MIN_LOTS = 2       # alt threshold for financial holdings

DEFAULT_PRIVACY_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    ("ssn",    re.compile(r"\b\d{3}-\d{2}-\d{4}\b")),
    ("card16", re.compile(r"\b(?:\d[ -]?){13,19}\b")),
    ("secret", re.compile(r"\b[A-Za-z0-9_/+=-]{32,}\b")),  # high-entropy string
)

REDACTED = "[REDACTED — privacy scan]"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

DUE_AMBER_FILL = PatternFill("solid", fgColor="FFE699")
DUE_RED_FILL = PatternFill("solid", fgColor="F4B084")
DUE_PAST_FONT = Font(bold=True, color="9C0006")


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class SheetSpec:
    """Declarative specification for one sheet within a workbook."""
    name: str
    columns: list[str]
    rows: list[list[Any]]
    column_widths: dict[int, int] = field(default_factory=dict)


@dataclass
class RenderResult:
    """Outcome of a single domain / entity render attempt."""
    target: str           # e.g. "domain:finances" or "entity:pet:buddy"
    path: Path | None     # path to written xlsx (None when skipped)
    status: str           # rendered | skipped-empty | skipped-stale | skipped-disabled | error
    sheets: list[str] = field(default_factory=list)
    redactions: int = 0
    error: str = ""


# ---------------------------------------------------------------------------
# IO helpers
# ---------------------------------------------------------------------------

def workspace_default(framework: Path) -> Path:
    return framework.parent / "workspace"


def load_yaml(path: Path) -> Any:
    if not path.exists():
        return None
    try:
        with path.open() as fh:
            return yaml.safe_load(fh)
    except (OSError, yaml.YAMLError):
        return None


def real_rows(data: Any, key: str) -> list[dict[str, Any]]:
    """Return rows under `data[key]` that are real entries (not template stubs).

    Every shipped memory template includes one example row with empty `id`
    (and other empty fields) as a schema reference. Without filtering those
    out, every workbook would contain a phantom row for an entity that
    doesn't exist. This helper enforces the convention.
    """
    if not isinstance(data, dict):
        return []
    rows = data.get(key) or []
    if not isinstance(rows, list):
        return []
    return [r for r in rows if isinstance(r, dict) and r.get("id")]


def now_iso() -> str:
    return dt.datetime.now().astimezone().isoformat(timespec="seconds")


def load_config(workspace: Path) -> dict[str, Any]:
    """Load `_memory/config.yaml`. Returns {} if missing."""
    return load_yaml(workspace / "_memory" / "config.yaml") or {}


def workbook_config(workspace: Path) -> dict[str, Any]:
    """Return preferences.workbooks block with sensible fallbacks."""
    cfg = load_config(workspace)
    wb_cfg: dict[str, Any] = (cfg.get("preferences") or {}).get("workbooks") or {}
    wb_cfg.setdefault("enabled", True)
    wb_cfg.setdefault("history_window_years", DEFAULT_HISTORY_YEARS)
    wb_cfg.setdefault("privacy_scan", "strict")
    wb_cfg.setdefault("per_entity", True)
    wb_cfg.setdefault("health", {"enabled": False})
    wb_cfg.setdefault("business", {"enabled": False})
    return wb_cfg


def write_config_workbooks(workspace: Path, mutate: Callable[[dict[str, Any]], None]) -> None:
    """Apply `mutate` to preferences.workbooks and write back to config.yaml."""
    config_path = workspace / "_memory" / "config.yaml"
    cfg = load_config(workspace)
    cfg.setdefault("preferences", {})
    cfg["preferences"].setdefault("workbooks", workbook_config(workspace))
    mutate(cfg["preferences"]["workbooks"])
    cfg["last_updated"] = now_iso()
    tmp = config_path.with_suffix(config_path.suffix + ".tmp")
    with tmp.open("w") as fh:
        yaml.safe_dump(cfg, fh, sort_keys=False, allow_unicode=True)
    tmp.replace(config_path)


# ---------------------------------------------------------------------------
# Privacy scan
# ---------------------------------------------------------------------------

def _sanitize_cell(value: Any, posture: str, redactions: list[int]) -> Any:
    """Apply the configured privacy posture to a single cell value."""
    if posture == "off" or value is None:
        return value
    text = str(value)
    if not text:
        return value
    for kind, pattern in DEFAULT_PRIVACY_PATTERNS:
        if posture == "lenient" and kind == "secret":
            continue
        if pattern.search(text):
            redactions[0] += 1
            return REDACTED
    return value


# ---------------------------------------------------------------------------
# Mtime / config sidecar
# ---------------------------------------------------------------------------

def _meta_path(workbook_path: Path) -> Path:
    return workbook_path.with_suffix(workbook_path.suffix + ".meta.yaml")


def _config_signature(cfg: dict[str, Any]) -> str:
    """Hash of the workbook-config bits that affect render output."""
    blob = (
        f"history_window_years={cfg.get('history_window_years')}|"
        f"privacy_scan={cfg.get('privacy_scan')}"
    )
    return hashlib.sha256(blob.encode()).hexdigest()[:16]


def needs_render(
    workbook_path: Path,
    source_paths: list[Path],
    cfg: dict[str, Any],
) -> bool:
    """True iff workbook is missing OR a source mtime / config has changed."""
    if not workbook_path.exists():
        return True
    meta = load_yaml(_meta_path(workbook_path)) or {}
    if meta.get("config_signature") != _config_signature(cfg):
        return True
    recorded = meta.get("source_mtimes") or {}
    for src in source_paths:
        if not src.exists():
            if recorded.get(str(src)) is not None:
                return True
            continue
        if str(src) not in recorded:
            return True
        if abs(recorded[str(src)] - src.stat().st_mtime) > 0.5:
            return True
    return False


def write_meta(
    workbook_path: Path,
    source_paths: list[Path],
    cfg: dict[str, Any],
) -> None:
    meta = {
        "schema_version": 1,
        "rendered_at": now_iso(),
        "config_signature": _config_signature(cfg),
        "source_mtimes": {
            str(p): p.stat().st_mtime for p in source_paths if p.exists()
        },
    }
    tmp = _meta_path(workbook_path).with_suffix(".tmp")
    with tmp.open("w") as fh:
        yaml.safe_dump(meta, fh, sort_keys=False, allow_unicode=True)
    tmp.replace(_meta_path(workbook_path))


# ---------------------------------------------------------------------------
# Sheet rendering
# ---------------------------------------------------------------------------

def _format_value(value: Any) -> Any:
    """Coerce YAML values into something openpyxl renders cleanly."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, (int, float, dt.date, dt.datetime)):
        return value
    if isinstance(value, list):
        return ", ".join(_format_value(v) for v in value if v not in (None, ""))
    if isinstance(value, dict):
        return ", ".join(f"{k}={_format_value(v)}" for k, v in value.items())
    return str(value)


def _render_sheet(wb: Workbook, spec: SheetSpec, posture: str, redactions: list[int]) -> None:
    ws = wb.create_sheet(spec.name)
    for col_idx, header in enumerate(spec.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(spec.columns))}1"
    )
    for row_idx, row in enumerate(spec.rows, start=2):
        for col_idx, raw in enumerate(row, start=1):
            value = _sanitize_cell(_format_value(raw), posture, redactions)
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx, header in enumerate(spec.columns, start=1):
        sample = [str(header)] + [
            str(_format_value(row[col_idx - 1]))
            for row in spec.rows[:50]
            if col_idx - 1 < len(row)
        ]
        width = max(len(s) for s in sample) + 2
        if col_idx in spec.column_widths:
            width = spec.column_widths[col_idx]
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 60)


def _build_workbook(specs: list[SheetSpec], posture: str) -> tuple[Workbook | None, int]:
    """Return (Workbook, redaction_count). None when no spec has any rows."""
    populated = [s for s in specs if s.rows]
    if not populated:
        return None, 0
    wb = Workbook()
    wb.remove(wb.active)
    redactions = [0]
    for spec in populated:
        _render_sheet(wb, spec, posture, redactions)
    return wb, redactions[0]


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def history_cutoff(cfg: dict[str, Any]) -> dt.date:
    years = int(cfg.get("history_window_years") or DEFAULT_HISTORY_YEARS)
    return dt.date.today() - dt.timedelta(days=365 * years)


def _parse_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.date):
        return value
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, str) and value:
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S%z"):
            try:
                return dt.datetime.strptime(value[: len(fmt) + 6], fmt).date()
            except ValueError:
                continue
        try:
            return dt.date.fromisoformat(value[:10])
        except ValueError:
            return None
    return None


def _within_window(value: Any, cutoff: dt.date) -> bool:
    d = _parse_date(value)
    return d is None or d >= cutoff


# ---------------------------------------------------------------------------
# Per-domain renderers
# ---------------------------------------------------------------------------

def _bills_rows(workspace: Path, cutoff: dt.date) -> list[SheetSpec]:
    data = load_yaml(workspace / "_memory" / "bills.yaml") or {}
    bills = real_rows(data, "bills")
    rows = [
        [
            r.get("name") or r.get("id"),
            r.get("kind"),
            r.get("amount"),
            r.get("currency", "USD"),
            r.get("cadence"),
            r.get("next_due"),
            r.get("pay_from_account") or r.get("pay_from"),
            r.get("status", "active"),
            r.get("autopay"),
            ", ".join(r.get("tags") or []),
        ]
        for r in bills
    ]
    if not rows:
        return []
    return [SheetSpec(
        name="Bills",
        columns=["Name", "Kind", "Amount", "Currency", "Cadence",
                 "Next due", "Pay from", "Status", "Auto-pay", "Tags"],
        rows=rows,
    )]


def _subscriptions_rows(workspace: Path) -> list[SheetSpec]:
    data = load_yaml(workspace / "_memory" / "subscriptions.yaml") or {}
    subs = real_rows(data, "subscriptions")
    rows = [
        [
            r.get("name") or r.get("id"),
            r.get("amount"),
            r.get("currency", "USD"),
            r.get("cadence"),
            r.get("last_used"),
            r.get("status", "active"),
            r.get("audit_flag"),
            ", ".join(r.get("tags") or []),
        ]
        for r in subs
    ]
    if not rows:
        return []
    return [SheetSpec(
        name="Subscriptions",
        columns=["Name", "Amount", "Currency", "Cadence", "Last used",
                 "Status", "Audit flag", "Tags"],
        rows=rows,
    )]


def _accounts_rows(workspace: Path) -> list[SheetSpec]:
    data = load_yaml(workspace / "_memory" / "accounts-index.yaml") or {}
    accts = real_rows(data, "accounts")
    rows = [
        [
            r.get("name") or r.get("id"),
            r.get("institution"),
            r.get("kind"),
            r.get("number_last4"),
            r.get("currency", "USD"),
            r.get("balance"),
            r.get("balance_as_of"),
            r.get("status", "active"),
        ]
        for r in accts
    ]
    if not rows:
        return []
    return [SheetSpec(
        name="Accounts",
        columns=["Name", "Institution", "Kind", "Last-4", "Currency",
                 "Balance (last known)", "As of", "Status"],
        rows=rows,
    )]


def _transactions_rows(workspace: Path, cutoff: dt.date) -> list[SheetSpec]:
    """Walk every account's transactions[] and surface the recent slice."""
    data = load_yaml(workspace / "_memory" / "accounts-index.yaml") or {}
    rows: list[list[Any]] = []
    tax_deductible: list[list[Any]] = []
    for acct in real_rows(data, "accounts"):
        acct_label = acct.get("name") or acct.get("id") or ""
        for txn in (acct.get("transactions") or []):
            if not isinstance(txn, dict):
                continue
            ts = txn.get("timestamp") or txn.get("date")
            if not _within_window(ts, cutoff):
                continue
            row = [
                ts,
                acct_label,
                txn.get("counterparty"),
                txn.get("amount"),
                txn.get("direction"),
                txn.get("channel"),
                txn.get("status"),
                txn.get("purpose"),
                txn.get("confirmation"),
                ", ".join(txn.get("related_entities") or []),
            ]
            rows.append(row)
            if "tax-deductible" in (txn.get("tags") or []):
                tax_deductible.append(row)
    rows.sort(key=lambda r: str(r[0] or ""), reverse=True)
    tax_deductible.sort(key=lambda r: str(r[0] or ""), reverse=True)
    specs: list[SheetSpec] = []
    if rows:
        specs.append(SheetSpec(
            name="Transactions",
            columns=["Timestamp", "Account", "Counterparty", "Amount",
                     "Direction", "Channel", "Status", "Purpose",
                     "Confirmation", "Related entities"],
            rows=rows,
        ))
    if tax_deductible:
        specs.append(SheetSpec(
            name="Tax-Deductible",
            columns=["Timestamp", "Account", "Counterparty", "Amount",
                     "Direction", "Channel", "Status", "Purpose",
                     "Confirmation", "Related entities"],
            rows=tax_deductible,
        ))
    return specs


def render_finances(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    cutoff = history_cutoff(cfg)
    specs = (
        _bills_rows(workspace, cutoff)
        + _subscriptions_rows(workspace)
        + _accounts_rows(workspace)
        + _transactions_rows(workspace, cutoff)
    )
    sources = [
        workspace / "_memory" / "bills.yaml",
        workspace / "_memory" / "subscriptions.yaml",
        workspace / "_memory" / "accounts-index.yaml",
    ]
    return specs, sources


def render_home(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    assets = load_yaml(workspace / "_memory" / "assets-index.yaml") or {}
    bills = load_yaml(workspace / "_memory" / "bills.yaml") or {}
    contacts = load_yaml(workspace / "_memory" / "contacts.yaml") or {}
    specs: list[SheetSpec] = []

    # Properties
    properties = [
        a for a in real_rows(assets, "assets")
        if a.get("kind") == "real_estate" and a.get("domain") == "home"
    ]
    if properties:
        specs.append(SheetSpec(
            name="Properties",
            columns=["Name", "Address", "APN", "County", "Acres",
                     "Purchase date", "Current value", "Insurance",
                     "Status"],
            rows=[
                [
                    a.get("name") or a.get("id"),
                    a.get("parcel_address") or a.get("location"),
                    a.get("apn"),
                    a.get("county"),
                    a.get("acres"),
                    a.get("purchase_date"),
                    a.get("current_value"),
                    a.get("insurance"),
                    a.get("status", "active"),
                ]
                for a in properties
            ],
        ))

    # Maintenance schedule across all home assets
    maint_rows: list[list[Any]] = []
    for a in real_rows(assets, "assets"):
        if a.get("domain") != "home":
            continue
        asset_label = a.get("name") or a.get("id")
        for m in (a.get("maintenance") or []):
            if not isinstance(m, dict):
                continue
            maint_rows.append([
                m.get("kind"),
                asset_label,
                m.get("interval_days"),
                m.get("last_done"),
                m.get("next_due"),
                m.get("vendor"),
            ])
    if maint_rows:
        specs.append(SheetSpec(
            name="Maintenance Schedule",
            columns=["Task", "Asset", "Interval (days)", "Last done",
                     "Next due", "Vendor"],
            rows=maint_rows,
        ))

    # Utilities filtered from bills
    util_rows = [
        [
            b.get("name") or b.get("id"),
            b.get("vendor"),
            b.get("number_last4"),
            b.get("amount"),
            b.get("cadence"),
            b.get("contact"),
        ]
        for b in real_rows(bills, "bills")
        if (b.get("related_domain") == "home" or b.get("kind") == "utility")
    ]
    if util_rows:
        specs.append(SheetSpec(
            name="Utilities",
            columns=["Name", "Vendor", "Last-4", "Amount",
                     "Cadence", "Contact"],
            rows=util_rows,
        ))

    # Contractors from contacts
    contractor_rows = [
        [
            c.get("name") or c.get("id"),
            c.get("relationship") or c.get("role"),
            c.get("phone"),
            c.get("email"),
            c.get("last_contacted"),
            c.get("notes"),
        ]
        for c in real_rows(contacts, "contacts")
        if "home" in (c.get("related_domains") or [])
    ]
    if contractor_rows:
        specs.append(SheetSpec(
            name="Contractors",
            columns=["Name", "Role", "Phone", "Email",
                     "Last contacted", "Notes"],
            rows=contractor_rows,
        ))

    sources = [
        workspace / "_memory" / "assets-index.yaml",
        workspace / "_memory" / "bills.yaml",
        workspace / "_memory" / "contacts.yaml",
    ]
    return specs, sources


def render_assets(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    data = load_yaml(workspace / "_memory" / "assets-index.yaml") or {}
    specs: list[SheetSpec] = []

    actives = [
        a for a in real_rows(data, "assets")
        if a.get("status", "active") == "active"
    ]

    physical_kinds = {"electronics", "appliance", "jewelry", "instrument",
                      "art", "collectible", "tool", "furniture", "sports_gear",
                      "other_physical"}
    financial_kinds = {"stock", "etf", "mutual_fund", "bond", "treasury",
                       "crypto", "cash_position", "precious_metal",
                       "other_financial"}

    physical = [a for a in actives if a.get("kind") in physical_kinds]
    financial = [a for a in actives if a.get("kind") in financial_kinds]
    realestate_other = [
        a for a in actives
        if a.get("kind") == "real_estate" and a.get("domain") != "home"
    ]

    if physical:
        specs.append(SheetSpec(
            name="Physical Inventory",
            columns=["Name", "Kind", "Sub-kind", "Make", "Model",
                     "Year", "Serial", "Location", "Purchase date",
                     "Purchase price", "Current value",
                     "Warranty expires"],
            rows=[
                [
                    a.get("name") or a.get("id"),
                    a.get("kind"),
                    a.get("sub_kind"),
                    a.get("make"),
                    a.get("model"),
                    a.get("year"),
                    a.get("serial"),
                    a.get("location"),
                    a.get("purchase_date"),
                    a.get("purchase_price"),
                    a.get("current_value"),
                    a.get("warranty_expires"),
                ]
                for a in physical
            ],
        ))

    if financial:
        specs.append(SheetSpec(
            name="Financial Holdings",
            columns=["Name", "Kind", "Ticker", "Exchange", "Units",
                     "Cost basis", "Current value", "Acquired",
                     "Held in account", "Notes"],
            rows=[
                [
                    a.get("name") or a.get("id"),
                    a.get("kind"),
                    a.get("ticker"),
                    a.get("exchange"),
                    a.get("units"),
                    a.get("cost_basis"),
                    a.get("current_value"),
                    a.get("acquired_at"),
                    a.get("held_in_account"),
                    a.get("notes"),
                ]
                for a in financial
            ],
        ))

    if realestate_other:
        specs.append(SheetSpec(
            name="Real Estate (non-residence)",
            columns=["Name", "Sub-kind", "Address", "APN", "Acres",
                     "Purchase date", "Current value", "Insurance",
                     "Status"],
            rows=[
                [
                    a.get("name") or a.get("id"),
                    a.get("sub_kind"),
                    a.get("parcel_address"),
                    a.get("apn"),
                    a.get("acres"),
                    a.get("purchase_date"),
                    a.get("current_value"),
                    a.get("insurance"),
                    a.get("status"),
                ]
                for a in realestate_other
            ],
        ))

    # Warranties — physical assets with a warranty expiration
    warr_rows = [
        [
            a.get("name") or a.get("id"),
            a.get("warranty_expires"),
            a.get("make"),
            a.get("model"),
            a.get("notes"),
        ]
        for a in physical
        if a.get("warranty_expires")
    ]
    if warr_rows:
        specs.append(SheetSpec(
            name="Warranties",
            columns=["Asset", "Expires", "Make", "Model", "Notes"],
            rows=warr_rows,
        ))

    # Disposals
    disposed = [a for a in real_rows(data, "assets")
                if a.get("status") == "disposed"]
    extra = data.get("disposed") or []
    disposed += [d for d in extra if isinstance(d, dict) and d.get("id")]
    if disposed:
        specs.append(SheetSpec(
            name="Disposals",
            columns=["Name", "Kind", "Date", "Reason", "Proceeds"],
            rows=[
                [
                    a.get("name") or a.get("id"),
                    a.get("kind"),
                    (a.get("disposed") or {}).get("date") if isinstance(a.get("disposed"), dict) else None,
                    (a.get("disposed") or {}).get("reason") if isinstance(a.get("disposed"), dict) else None,
                    (a.get("disposed") or {}).get("proceeds") if isinstance(a.get("disposed"), dict) else None,
                ]
                for a in disposed
            ],
        ))

    sources = [
        workspace / "_memory" / "assets-index.yaml",
        workspace / "_memory" / "accounts-index.yaml",
    ]
    return specs, sources


def render_health(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    if not (cfg.get("health") or {}).get("enabled", False):
        return [], []
    data = load_yaml(workspace / "_memory" / "health-records.yaml") or {}
    cutoff = history_cutoff(cfg)
    specs: list[SheetSpec] = []
    if data.get("vitals"):
        specs.append(SheetSpec(
            name="Vitals",
            columns=["Date", "Type", "Value", "Unit", "Source"],
            rows=[
                [v.get("date"), v.get("type"), v.get("value"),
                 v.get("unit"), v.get("source")]
                for v in data["vitals"]
                if isinstance(v, dict) and _within_window(v.get("date"), cutoff)
            ],
        ))
    if data.get("medications"):
        specs.append(SheetSpec(
            name="Medications",
            columns=["Name", "Dose", "Schedule", "Prescriber",
                     "Started", "Stopped", "Status"],
            rows=[
                [m.get("name"), m.get("dose"), m.get("schedule"),
                 m.get("prescriber"), m.get("started"), m.get("stopped"),
                 m.get("status")]
                for m in data["medications"] if isinstance(m, dict)
            ],
        ))
    if data.get("vaccines"):
        specs.append(SheetSpec(
            name="Vaccines",
            columns=["Date", "Vaccine", "Dose #", "Lot", "Provider",
                     "Next due"],
            rows=[
                [v.get("date"), v.get("vaccine"), v.get("dose_number"),
                 v.get("lot"), v.get("provider"), v.get("next_due")]
                for v in data["vaccines"]
                if isinstance(v, dict) and _within_window(v.get("date"), cutoff)
            ],
        ))
    if data.get("visits"):
        specs.append(SheetSpec(
            name="Visits",
            columns=["Date", "Provider", "Kind", "Reason", "Outcome",
                     "Follow-up"],
            rows=[
                [v.get("date"), v.get("provider"), v.get("kind"),
                 v.get("reason"), v.get("outcome"), v.get("follow_up")]
                for v in data["visits"]
                if isinstance(v, dict) and _within_window(v.get("date"), cutoff)
            ],
        ))
    if data.get("results"):
        specs.append(SheetSpec(
            name="Lab Results",
            columns=["Date", "Panel", "Summary", "Ranges", "Doctor's read",
                     "File ref"],
            rows=[
                [r.get("date"), r.get("panel"), r.get("summary"),
                 r.get("ranges"), r.get("doctor_read"), r.get("file_ref")]
                for r in data["results"]
                if isinstance(r, dict) and _within_window(r.get("date"), cutoff)
            ],
        ))
    if data.get("symptoms"):
        specs.append(SheetSpec(
            name="Symptoms",
            columns=["Date", "Symptom", "Severity", "Context", "Resolved"],
            rows=[
                [s.get("date"), s.get("symptom"), s.get("severity"),
                 s.get("context"), s.get("resolved")]
                for s in data["symptoms"]
                if isinstance(s, dict) and _within_window(s.get("date"), cutoff)
            ],
        ))
    sources = [workspace / "_memory" / "health-records.yaml"]
    return specs, sources


def render_vehicles(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    data = load_yaml(workspace / "_memory" / "assets-index.yaml") or {}
    vehicles = [a for a in real_rows(data, "assets") if a.get("domain") == "vehicles"]
    specs: list[SheetSpec] = []
    if vehicles:
        specs.append(SheetSpec(
            name="Vehicles",
            columns=["Name", "Make", "Model", "Year", "VIN (last 4)",
                     "Color", "Registration expires", "Insurance carrier",
                     "Mileage", "Status"],
            rows=[
                [
                    a.get("name") or a.get("id"),
                    a.get("make"),
                    a.get("model"),
                    a.get("year"),
                    (a.get("vin") or "")[-4:] if a.get("vin") else "",
                    a.get("color"),
                    a.get("registration_expires"),
                    a.get("insurance"),
                    a.get("current_mileage"),
                    a.get("status", "active"),
                ]
                for a in vehicles
            ],
        ))
        # Service history aggregated across all vehicles
        srv_rows: list[list[Any]] = []
        for v in vehicles:
            v_label = v.get("name") or v.get("id")
            for m in (v.get("maintenance") or []):
                if not isinstance(m, dict):
                    continue
                srv_rows.append([
                    m.get("last_done") or m.get("next_due"),
                    v_label,
                    m.get("kind"),
                    m.get("interval_days"),
                    m.get("interval_miles"),
                    m.get("next_due"),
                    m.get("vendor"),
                ])
        if srv_rows:
            specs.append(SheetSpec(
                name="Service History",
                columns=["Date", "Vehicle", "Service", "Interval (days)",
                         "Interval (miles)", "Next due", "Vendor"],
                rows=srv_rows,
            ))
    sources = [workspace / "_memory" / "assets-index.yaml"]
    return specs, sources


def render_pets(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    contacts = load_yaml(workspace / "_memory" / "contacts.yaml") or {}
    appts = load_yaml(workspace / "_memory" / "appointments.yaml") or {}
    pets = [
        c for c in real_rows(contacts, "contacts")
        if c.get("role") == "pet" or "pets" in (c.get("related_domains") or [])
    ]
    specs: list[SheetSpec] = []
    if pets:
        specs.append(SheetSpec(
            name="Pets",
            columns=["Name", "Species / Breed", "DOB", "Microchip",
                     "Notes"],
            rows=[
                [
                    p.get("name"),
                    p.get("relationship") or p.get("notes"),
                    p.get("dob"),
                    p.get("microchip_last4"),
                    p.get("notes"),
                ]
                for p in pets
            ],
        ))
        cutoff = history_cutoff(cfg)
        visit_rows = [
            [
                a.get("date"),
                a.get("for_member"),
                a.get("provider"),
                a.get("reason"),
                a.get("outcome"),
            ]
            for a in real_rows(appts, "appointments")
            if a.get("kind") == "vet" and _within_window(a.get("date"), cutoff)
        ]
        if visit_rows:
            specs.append(SheetSpec(
                name="Vet Visits",
                columns=["Date", "Pet", "Vet", "Reason", "Outcome"],
                rows=visit_rows,
            ))
    sources = [
        workspace / "_memory" / "contacts.yaml",
        workspace / "_memory" / "appointments.yaml",
    ]
    return specs, sources


def render_family(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    contacts = load_yaml(workspace / "_memory" / "contacts.yaml") or {}
    dates = load_yaml(workspace / "_memory" / "important-dates.yaml") or {}
    family_roles = {"family", "spouse", "partner", "parent", "child",
                    "sibling", "in-law"}
    specs: list[SheetSpec] = []
    members = [
        c for c in real_rows(contacts, "contacts")
        if c.get("role") in family_roles
        or c.get("relationship") in family_roles
        or "family" in (c.get("related_domains") or [])
    ]
    if members:
        specs.append(SheetSpec(
            name="Members",
            columns=["Name", "Relationship", "DOB", "Phone", "Email",
                     "Notes"],
            rows=[
                [c.get("name"), c.get("relationship"), c.get("dob"),
                 c.get("phone"), c.get("email"), c.get("notes")]
                for c in members
            ],
        ))
    date_rows = real_rows(dates, "dates") or real_rows(dates, "important_dates")
    bday_rows = [
        [d.get("next_occurrence") or d.get("date"),
         d.get("title"), d.get("kind"), d.get("recurrence")]
        for d in date_rows
        if d.get("related_domain") == "family"
        or d.get("kind") in {"birthday", "anniversary"}
    ]
    if bday_rows:
        specs.append(SheetSpec(
            name="Birthdays & Anniversaries",
            columns=["Next occurrence", "Person", "Kind", "Recurrence"],
            rows=bday_rows,
        ))
    sources = [
        workspace / "_memory" / "contacts.yaml",
        workspace / "_memory" / "important-dates.yaml",
    ]
    return specs, sources


def render_travel(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    docs = load_yaml(workspace / "_memory" / "documents-index.yaml") or {}
    appts = load_yaml(workspace / "_memory" / "appointments.yaml") or {}
    specs: list[SheetSpec] = []
    expirations = [
        d for d in real_rows(docs, "documents")
        if d.get("related_domain") == "travel"
    ]
    if expirations:
        specs.append(SheetSpec(
            name="Document Expirations",
            columns=["Title", "Kind", "Expires", "For member", "Notes"],
            rows=[
                [d.get("title"), d.get("kind"), d.get("expires_at"),
                 d.get("for_member"), d.get("notes")]
                for d in expirations
            ],
        ))
    cutoff = history_cutoff(cfg)
    bookings = [
        [a.get("date"), a.get("kind"), a.get("provider"),
         a.get("title"), a.get("confirmation")]
        for a in real_rows(appts, "appointments")
        if a.get("kind") in {"travel", "flight", "hotel", "rental_car"}
        and _within_window(a.get("date"), cutoff)
    ]
    if bookings:
        specs.append(SheetSpec(
            name="Bookings",
            columns=["Date", "Kind", "Provider", "Title", "Confirmation"],
            rows=bookings,
        ))
    sources = [
        workspace / "_memory" / "documents-index.yaml",
        workspace / "_memory" / "appointments.yaml",
    ]
    return specs, sources


def render_career(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    docs = load_yaml(workspace / "_memory" / "documents-index.yaml") or {}
    contacts = load_yaml(workspace / "_memory" / "contacts.yaml") or {}
    specs: list[SheetSpec] = []
    certs = [
        d for d in real_rows(docs, "documents")
        if d.get("kind") in {"certification", "license"}
        and d.get("related_domain") == "career"
    ]
    if certs:
        specs.append(SheetSpec(
            name="Certifications",
            columns=["Title", "Issuer", "Earned", "Expires", "Status",
                     "Renewal cost"],
            rows=[
                [d.get("title"), d.get("issuer"), d.get("issued_at"),
                 d.get("expires_at"), d.get("status"),
                 d.get("renewal_cost")]
                for d in certs
            ],
        ))
    network = [
        c for c in real_rows(contacts, "contacts")
        if "career" in (c.get("related_domains") or [])
    ]
    if network:
        specs.append(SheetSpec(
            name="Network",
            columns=["Name", "Organization", "Role", "Last contacted",
                     "Notes"],
            rows=[
                [c.get("name"), c.get("organization"),
                 c.get("relationship") or c.get("role"),
                 c.get("last_contacted"), c.get("notes")]
                for c in network
            ],
        ))
    sources = [
        workspace / "_memory" / "documents-index.yaml",
        workspace / "_memory" / "contacts.yaml",
    ]
    return specs, sources


def render_business(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    if not (cfg.get("business") or {}).get("enabled", False):
        return [], []
    contacts = load_yaml(workspace / "_memory" / "contacts.yaml") or {}
    invoices = load_yaml(workspace / "_memory" / "invoices.yaml") or {}
    docs = load_yaml(workspace / "_memory" / "documents-index.yaml") or {}
    accts = load_yaml(workspace / "_memory" / "accounts-index.yaml") or {}
    specs: list[SheetSpec] = []
    clients = [
        c for c in real_rows(contacts, "contacts")
        if "business" in (c.get("related_domains") or [])
    ]
    if clients:
        specs.append(SheetSpec(
            name="Clients",
            columns=["Name", "Organization", "Status", "Last contacted",
                     "Notes"],
            rows=[
                [c.get("name"), c.get("organization"),
                 c.get("status", "active"),
                 c.get("last_contacted"), c.get("notes")]
                for c in clients
            ],
        ))
    inv_rows = [
        [
            i.get("id"), i.get("client"), i.get("issue_date"),
            i.get("due_date"), i.get("amount"), i.get("currency", "USD"),
            i.get("status"), i.get("paid_date"), i.get("paid_amount"),
            i.get("payment_confirmation"),
        ]
        for i in real_rows(invoices, "invoices")
    ]
    if inv_rows:
        specs.append(SheetSpec(
            name="Invoices",
            columns=["Invoice #", "Client", "Issued", "Due", "Amount",
                     "Currency", "Status", "Paid date", "Paid amount",
                     "Confirmation"],
            rows=inv_rows,
        ))
    contracts = [
        d for d in real_rows(docs, "documents")
        if d.get("kind") == "contract"
        and d.get("related_domain") == "business"
    ]
    if contracts:
        specs.append(SheetSpec(
            name="Contracts",
            columns=["Title", "Client", "Signed", "Term ends", "Value",
                     "Status"],
            rows=[
                [d.get("title"), d.get("counterparty"), d.get("signed_at"),
                 d.get("expires_at"), d.get("value"), d.get("status")]
                for d in contracts
            ],
        ))
    cutoff = history_cutoff(cfg)
    biz_acct_ids = {
        a.get("id") for a in real_rows(accts, "accounts")
        if a.get("related_domain") == "business"
    }
    expense_rows: list[list[Any]] = []
    for a in real_rows(accts, "accounts"):
        if a.get("id") not in biz_acct_ids:
            continue
        for txn in (a.get("transactions") or []):
            if not isinstance(txn, dict):
                continue
            if txn.get("direction") not in {"out", "fee"}:
                continue
            if not _within_window(txn.get("timestamp") or txn.get("date"), cutoff):
                continue
            expense_rows.append([
                txn.get("timestamp") or txn.get("date"),
                txn.get("counterparty"),
                txn.get("amount"),
                txn.get("purpose"),
                "yes" if "tax-deductible" in (txn.get("tags") or []) else "",
            ])
    if expense_rows:
        specs.append(SheetSpec(
            name="Expenses",
            columns=["Date", "Vendor", "Amount", "Purpose", "Deductible"],
            rows=expense_rows,
        ))
    sources = [
        workspace / "_memory" / "contacts.yaml",
        workspace / "_memory" / "invoices.yaml",
        workspace / "_memory" / "documents-index.yaml",
        workspace / "_memory" / "accounts-index.yaml",
    ]
    return specs, sources


def render_education(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    docs = load_yaml(workspace / "_memory" / "documents-index.yaml") or {}
    contacts = load_yaml(workspace / "_memory" / "contacts.yaml") or {}
    dates = load_yaml(workspace / "_memory" / "important-dates.yaml") or {}
    specs: list[SheetSpec] = []
    edu_docs = [
        d for d in real_rows(docs, "documents")
        if d.get("related_domain") == "education"
    ]
    if edu_docs:
        specs.append(SheetSpec(
            name="Documents",
            columns=["Title", "Kind", "Issued", "Expires", "Issuer",
                     "Notes"],
            rows=[
                [d.get("title"), d.get("kind"), d.get("issued_at"),
                 d.get("expires_at"), d.get("issuer"), d.get("notes")]
                for d in edu_docs
            ],
        ))
    advisors = [
        c for c in real_rows(contacts, "contacts")
        if "education" in (c.get("related_domains") or [])
    ]
    if advisors:
        specs.append(SheetSpec(
            name="Advisors",
            columns=["Name", "Organization", "Role", "Phone", "Email",
                     "Last contacted"],
            rows=[
                [c.get("name"), c.get("organization"),
                 c.get("relationship") or c.get("role"),
                 c.get("phone"), c.get("email"), c.get("last_contacted")]
                for c in advisors
            ],
        ))
    date_rows = real_rows(dates, "dates") or real_rows(dates, "important_dates")
    deadlines = [d for d in date_rows if d.get("related_domain") == "education"]
    if deadlines:
        specs.append(SheetSpec(
            name="Deadlines",
            columns=["Date", "Title", "Kind", "Recurrence"],
            rows=[
                [d.get("next_occurrence") or d.get("date"), d.get("title"),
                 d.get("kind"), d.get("recurrence")]
                for d in deadlines
            ],
        ))
    sources = [
        workspace / "_memory" / "documents-index.yaml",
        workspace / "_memory" / "contacts.yaml",
        workspace / "_memory" / "important-dates.yaml",
    ]
    return specs, sources


def render_hobbies(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    todos = load_yaml(workspace / "_memory" / "todo.yaml") or {}
    specs: list[SheetSpec] = []
    goals = [
        t for t in real_rows(todos, "tasks")
        if t.get("related_domain") == "hobbies"
    ]
    if goals:
        specs.append(SheetSpec(
            name="Goals & Open Tasks",
            columns=["Priority", "Title", "Status", "Due", "Created"],
            rows=[
                [t.get("priority"), t.get("title"), t.get("status"),
                 t.get("due"), t.get("created")]
                for t in goals
            ],
        ))
    sources = [workspace / "_memory" / "todo.yaml"]
    return specs, sources


def render_self(workspace: Path, cfg: dict[str, Any]) -> tuple[list[SheetSpec], list[Path]]:
    signals = load_yaml(workspace / "_memory" / "personal-signals.yaml") or {}
    decisions = load_yaml(workspace / "_memory" / "decisions.yaml") or {}
    specs: list[SheetSpec] = []
    cutoff = history_cutoff(cfg)
    sig_rows = [
        [s.get("captured_at") or s.get("at"), s.get("category") or s.get("theme"),
         s.get("trigger_phrase") or s.get("signal") or s.get("note"),
         s.get("surfaced_count")]
        for s in (signals.get("signals") or [])
        if isinstance(s, dict)
        and s.get("id")  # template ships with empty-id stub
        and _within_window(s.get("captured_at") or s.get("at"), cutoff)
    ]
    if sig_rows:
        specs.append(SheetSpec(
            name="Personal Signals",
            columns=["Captured", "Theme", "Signal", "Surfaced count"],
            rows=sig_rows,
        ))
    dec_rows = [
        [d.get("ts") or d.get("at") or d.get("date"), d.get("decision"),
         d.get("rationale") or d.get("reasoning"),
         d.get("review_at"), d.get("outcome") or d.get("status")]
        for d in (decisions.get("decisions") or [])
        if isinstance(d, dict)
        and d.get("id") and d.get("decision")  # template stub has both empty
        and _within_window(d.get("ts") or d.get("at") or d.get("date"), cutoff)
    ]
    if dec_rows:
        specs.append(SheetSpec(
            name="Decisions",
            columns=["Date", "Decision", "Rationale", "Review at",
                     "Outcome"],
            rows=dec_rows,
        ))
    sources = [
        workspace / "_memory" / "personal-signals.yaml",
        workspace / "_memory" / "decisions.yaml",
    ]
    return specs, sources


DOMAIN_RENDERERS: dict[str, Callable[[Path, dict[str, Any]], tuple[list[SheetSpec], list[Path]]]] = {
    "finances": render_finances,
    "home": render_home,
    "assets": render_assets,
    "health": render_health,
    "vehicles": render_vehicles,
    "pets": render_pets,
    "family": render_family,
    "travel": render_travel,
    "career": render_career,
    "business": render_business,
    "education": render_education,
    "hobbies": render_hobbies,
    "self": render_self,
}


# ---------------------------------------------------------------------------
# Per-entity renderers
# ---------------------------------------------------------------------------

def _entity_event_count(workspace: Path, kind: str, slug: str) -> int:
    if kind == "pet":
        appts = load_yaml(workspace / "_memory" / "appointments.yaml") or {}
        return sum(
            1 for a in real_rows(appts, "appointments")
            if a.get("kind") == "vet"
            and (a.get("for_member") or "").endswith(slug)
        )
    if kind == "vehicle":
        assets = load_yaml(workspace / "_memory" / "assets-index.yaml") or {}
        for a in real_rows(assets, "assets"):
            if a.get("id") == slug:
                return len(a.get("maintenance") or [])
        return 0
    if kind == "asset":
        assets = load_yaml(workspace / "_memory" / "assets-index.yaml") or {}
        for a in real_rows(assets, "assets"):
            if a.get("id") == slug:
                return max(len(a.get("lots") or []), len(a.get("receipts") or []))
        return 0
    return 0


def render_pet_entity(workspace: Path, cfg: dict[str, Any], slug: str) -> tuple[list[SheetSpec], list[Path]]:
    appts = load_yaml(workspace / "_memory" / "appointments.yaml") or {}
    cutoff = history_cutoff(cfg)
    visits = [
        a for a in real_rows(appts, "appointments")
        if a.get("kind") == "vet"
        and (a.get("for_member") or "").endswith(slug)
        and _within_window(a.get("date"), cutoff)
    ]
    specs: list[SheetSpec] = []
    if visits:
        specs.append(SheetSpec(
            name="Vet Visits",
            columns=["Date", "Vet", "Reason", "Outcome", "Cost"],
            rows=[
                [v.get("date"), v.get("provider"), v.get("reason"),
                 v.get("outcome"), v.get("cost")]
                for v in visits
            ],
        ))
    sources = [workspace / "_memory" / "appointments.yaml"]
    return specs, sources


def render_vehicle_entity(workspace: Path, cfg: dict[str, Any], slug: str) -> tuple[list[SheetSpec], list[Path]]:
    data = load_yaml(workspace / "_memory" / "assets-index.yaml") or {}
    veh = next(
        (a for a in real_rows(data, "assets") if a.get("id") == slug),
        None,
    )
    specs: list[SheetSpec] = []
    if veh:
        maint = veh.get("maintenance") or []
        if maint:
            specs.append(SheetSpec(
                name="Service",
                columns=["Last done", "Service", "Interval (days)",
                         "Interval (miles)", "Next due", "Vendor"],
                rows=[
                    [m.get("last_done"), m.get("kind"),
                     m.get("interval_days"), m.get("interval_miles"),
                     m.get("next_due"), m.get("vendor")]
                    for m in maint if isinstance(m, dict)
                ],
            ))
    sources = [workspace / "_memory" / "assets-index.yaml"]
    return specs, sources


def render_asset_entity(workspace: Path, cfg: dict[str, Any], slug: str) -> tuple[list[SheetSpec], list[Path]]:
    data = load_yaml(workspace / "_memory" / "assets-index.yaml") or {}
    asset = next(
        (a for a in real_rows(data, "assets") if a.get("id") == slug),
        None,
    )
    specs: list[SheetSpec] = []
    if asset:
        lots = asset.get("lots") or []
        if lots:
            specs.append(SheetSpec(
                name="Lots",
                columns=["Acquired", "Units", "Cost basis", "Source"],
                rows=[
                    [lot.get("acquired_at"), lot.get("units"),
                     lot.get("cost_basis"), lot.get("source")]
                    for lot in lots if isinstance(lot, dict)
                ],
            ))
        receipts = asset.get("receipts") or []
        if receipts:
            specs.append(SheetSpec(
                name="Receipts",
                columns=["Path"],
                rows=[[r] for r in receipts],
            ))
    sources = [workspace / "_memory" / "assets-index.yaml"]
    return specs, sources


ENTITY_RENDERERS: dict[str, tuple[str, Callable[[Path, dict[str, Any], str], tuple[list[SheetSpec], list[Path]]]]] = {
    "pet": ("Pets", render_pet_entity),
    "vehicle": ("Vehicles", render_vehicle_entity),
    "asset": ("Assets", render_asset_entity),
}


# ---------------------------------------------------------------------------
# Top-level dispatch
# ---------------------------------------------------------------------------

def _ensure_domain_folder(workspace: Path, domain_name: str) -> None:
    """Materialize Domains/<Name>/ on demand (lazy-folder contract)."""
    try:
        from superagent.tools.domains import ensure_folder
    except ImportError:
        (workspace / "Domains" / domain_name).mkdir(parents=True, exist_ok=True)
        return
    framework = Path(__file__).resolve().parent.parent
    try:
        ensure_folder(workspace, framework, domain_name.lower())
    except ValueError:
        (workspace / "Domains" / domain_name).mkdir(parents=True, exist_ok=True)


def _domain_name_for(domain_id: str) -> str:
    return "Self" if domain_id == "self" else domain_id.title()


def render_domain(
    workspace: Path,
    framework: Path,
    domain_id: str,
    *,
    check: bool = False,
) -> RenderResult:
    cfg = workbook_config(workspace)
    if not cfg.get("enabled", True):
        return RenderResult(target=f"domain:{domain_id}", path=None,
                            status="skipped-disabled")
    renderer = DOMAIN_RENDERERS.get(domain_id)
    if renderer is None:
        return RenderResult(target=f"domain:{domain_id}", path=None,
                            status="error", error=f"unknown domain {domain_id!r}")
    specs, sources = renderer(workspace, cfg)
    domain_name = _domain_name_for(domain_id)
    workbook_path = workspace / "Domains" / domain_name / f"{domain_id}.xlsx"
    if not specs:
        # Lazy: don't emit empty workbook. Remove a stale one if present.
        if workbook_path.exists() and not check:
            workbook_path.unlink()
            meta = _meta_path(workbook_path)
            if meta.exists():
                meta.unlink()
        return RenderResult(target=f"domain:{domain_id}", path=None,
                            status="skipped-empty")
    if not needs_render(workbook_path, sources, cfg):
        return RenderResult(target=f"domain:{domain_id}", path=workbook_path,
                            status="skipped-stale",
                            sheets=[s.name for s in specs])
    if check:
        return RenderResult(target=f"domain:{domain_id}", path=workbook_path,
                            status="rendered", sheets=[s.name for s in specs])
    posture = cfg.get("privacy_scan", "strict")
    wb, redactions = _build_workbook(specs, posture)
    if wb is None:
        return RenderResult(target=f"domain:{domain_id}", path=None,
                            status="skipped-empty")
    _ensure_domain_folder(workspace, domain_name)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)
    write_meta(workbook_path, sources, cfg)
    return RenderResult(
        target=f"domain:{domain_id}",
        path=workbook_path,
        status="rendered",
        sheets=[s.name for s in specs],
        redactions=redactions,
    )


def render_entity(
    workspace: Path,
    framework: Path,
    kind: str,
    slug: str,
    *,
    check: bool = False,
) -> RenderResult:
    cfg = workbook_config(workspace)
    if not cfg.get("enabled", True):
        return RenderResult(target=f"entity:{kind}:{slug}", path=None,
                            status="skipped-disabled")
    if not cfg.get("per_entity", True):
        return RenderResult(target=f"entity:{kind}:{slug}", path=None,
                            status="skipped-disabled")
    if kind not in ENTITY_RENDERERS:
        return RenderResult(target=f"entity:{kind}:{slug}", path=None,
                            status="error", error=f"unknown entity kind {kind!r}")
    domain_name, renderer = ENTITY_RENDERERS[kind]
    events = _entity_event_count(workspace, kind, slug)
    threshold = (
        DEFAULT_PER_ENTITY_MIN_LOTS if kind == "asset"
        else DEFAULT_PER_ENTITY_MIN_EVENTS
    )
    if events < threshold:
        return RenderResult(target=f"entity:{kind}:{slug}", path=None,
                            status="skipped-empty")
    specs, sources = renderer(workspace, cfg, slug)
    workbook_path = workspace / "Domains" / domain_name / f"{slug}.xlsx"
    if not specs:
        if workbook_path.exists() and not check:
            workbook_path.unlink()
            meta = _meta_path(workbook_path)
            if meta.exists():
                meta.unlink()
        return RenderResult(target=f"entity:{kind}:{slug}", path=None,
                            status="skipped-empty")
    if not needs_render(workbook_path, sources, cfg):
        return RenderResult(target=f"entity:{kind}:{slug}", path=workbook_path,
                            status="skipped-stale",
                            sheets=[s.name for s in specs])
    if check:
        return RenderResult(target=f"entity:{kind}:{slug}", path=workbook_path,
                            status="rendered",
                            sheets=[s.name for s in specs])
    posture = cfg.get("privacy_scan", "strict")
    wb, redactions = _build_workbook(specs, posture)
    if wb is None:
        return RenderResult(target=f"entity:{kind}:{slug}", path=None,
                            status="skipped-empty")
    _ensure_domain_folder(workspace, domain_name)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)
    write_meta(workbook_path, sources, cfg)
    return RenderResult(
        target=f"entity:{kind}:{slug}",
        path=workbook_path,
        status="rendered",
        sheets=[s.name for s in specs],
        redactions=redactions,
    )


def render_all(workspace: Path, framework: Path, *, check: bool = False) -> list[RenderResult]:
    results: list[RenderResult] = []
    for domain_id in DOMAIN_RENDERERS:
        results.append(render_domain(workspace, framework, domain_id, check=check))
    cfg = workbook_config(workspace)
    if cfg.get("per_entity", True):
        # Pets
        contacts = load_yaml(workspace / "_memory" / "contacts.yaml") or {}
        for c in real_rows(contacts, "contacts"):
            if c.get("role") == "pet":
                results.append(render_entity(workspace, framework, "pet",
                                             c["id"], check=check))
        # Vehicles + Assets
        assets = load_yaml(workspace / "_memory" / "assets-index.yaml") or {}
        for a in real_rows(assets, "assets"):
            if a.get("domain") == "vehicles":
                results.append(render_entity(workspace, framework, "vehicle",
                                             a["id"], check=check))
            elif a.get("domain") == "assets":
                results.append(render_entity(workspace, framework, "asset",
                                             a["id"], check=check))
    return results


def status(workspace: Path) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    cfg = workbook_config(workspace)
    for domain_id in DOMAIN_RENDERERS:
        domain_name = _domain_name_for(domain_id)
        path = workspace / "Domains" / domain_name / f"{domain_id}.xlsx"
        out.append({
            "target": f"domain:{domain_id}",
            "path": str(path),
            "exists": path.exists(),
            "size_bytes": path.stat().st_size if path.exists() else 0,
            "rendered_at": (load_yaml(_meta_path(path)) or {}).get("rendered_at"),
            "opt_in_status": (
                "opt-in"
                if domain_id in {"health", "business"}
                and not (cfg.get(domain_id) or {}).get("enabled", False)
                else "default"
            ),
        })
    return out


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(prog="render_workbooks")
    parser.add_argument("--workspace", type=Path, default=None)
    parser.add_argument(
        "--framework",
        type=Path,
        default=Path(__file__).resolve().parent.parent,
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--domain", type=str, help="Render one domain workbook (e.g. finances).")
    group.add_argument(
        "--entity",
        type=str,
        help="Render one entity workbook in `<kind>:<slug>` form (e.g. pet:buddy).",
    )
    group.add_argument("--all", action="store_true", help="Render every applicable workbook.")
    group.add_argument("--status", action="store_true", help="Show per-workbook status.")
    group.add_argument("--enable", choices=("health", "business"))
    group.add_argument("--disable", choices=("health", "business"))
    parser.add_argument("--check", action="store_true",
                        help="Dry-run; report what WOULD render without writing.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    framework: Path = args.framework
    workspace: Path = args.workspace or workspace_default(framework)
    if not (workspace / "_memory").exists():
        print(f"no workspace at {workspace}", file=sys.stderr)
        return 1

    if args.enable:
        write_config_workbooks(
            workspace,
            lambda wb: wb.setdefault(args.enable, {}).update({"enabled": True}),
        )
        print(f"enabled {args.enable} workbook in config.preferences.workbooks.{args.enable}.enabled")
        return 0

    if args.disable:
        write_config_workbooks(
            workspace,
            lambda wb: wb.setdefault(args.disable, {}).update({"enabled": False}),
        )
        print(f"disabled {args.disable} workbook in config.preferences.workbooks.{args.disable}.enabled")
        return 0

    if args.status:
        for row in status(workspace):
            mark = "[ok]" if row["exists"] else "    "
            print(f"  {mark} {row['target']:25s} size={row['size_bytes']:>7d}b "
                  f"rendered={row['rendered_at'] or '—':25s} {row['opt_in_status']}")
        return 0

    results: list[RenderResult] = []
    if args.domain:
        results.append(render_domain(workspace, framework, args.domain, check=args.check))
    elif args.entity:
        if ":" not in args.entity:
            print("--entity must be `<kind>:<slug>` (e.g. pet:buddy)", file=sys.stderr)
            return 1
        kind, slug = args.entity.split(":", 1)
        results.append(render_entity(workspace, framework, kind, slug, check=args.check))
    elif args.all:
        results = render_all(workspace, framework, check=args.check)

    rendered = [r for r in results if r.status == "rendered"]
    skipped = [r for r in results if r.status.startswith("skipped")]
    errors = [r for r in results if r.status == "error"]
    for r in results:
        verb = "would render" if args.check and r.status == "rendered" else r.status
        if r.status == "rendered" or r.status == "skipped-stale":
            sheets = ", ".join(r.sheets) if r.sheets else "—"
            redactions = f" (redacted {r.redactions} cell(s))" if r.redactions else ""
            print(f"  {verb:18s} {r.target:25s} sheets: {sheets}{redactions}")
        elif r.status == "error":
            print(f"  ERROR             {r.target}: {r.error}", file=sys.stderr)
        else:
            print(f"  {verb:18s} {r.target}")
    print(f"\nSummary: rendered={len(rendered)} skipped={len(skipped)} errors={len(errors)}")
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
